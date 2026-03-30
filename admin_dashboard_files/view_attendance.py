import customtkinter as ctk
from PIL import Image
import os
from admin_dashboard_files.shared import present_details_list, present_student_ids, get_connection
from db import get_attendance_history



# -------------------- Present Students --------------------
def show_present_list_content(content_area, responsive_manager, data=None):

    """Displays a list of students currently recognized in this session."""
    from tkinter import filedialog, messagebox
    from datetime import datetime

    ctk.CTkLabel(content_area, text="✅ Recognized Students (Live)", font=("Segoe UI", 28, "bold"), text_color="#00c853").pack(anchor="w", padx=30, pady=(30, 10))
    summary_bar = ctk.CTkFrame(content_area, fg_color="#1a1a1a", corner_radius=10)
    summary_bar.pack(fill="x", padx=30, pady=(0, 20))
    ctk.CTkLabel(summary_bar, text=f"Total Recognized: {len(present_details_list)}", font=("Arial", 14, "bold"), text_color="gray").pack(side="left", padx=20, pady=10)

    # Removed local export button as requested.
    
    # -------------------- Present Table --------------------
    search_frame = ctk.CTkFrame(content_area, fg_color="transparent")
    search_frame.pack(fill="x", padx=30, pady=(0, 10))
    ctk.CTkLabel(search_frame, text="🔍 Search Student:", font=("Arial", 12, "bold"), text_color="gray").pack(side="left")
    search_entry = ctk.CTkEntry(search_frame, placeholder_text="Type Name, Reg No...", width=350, height=35, corner_radius=10)
    search_entry.pack(side="left", padx=15)

    table_frame = ctk.CTkFrame(content_area, fg_color="#1a1a1a", corner_radius=12)
    table_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
    header = ctk.CTkFrame(table_frame, fg_color="#0f0f0f", corner_radius=10)
    header.pack(fill="x", padx=15, pady=15)
    cols = ["Name", "Reg No", "Course", "Program", "Date", "Time"]
    for text in cols:
        ctk.CTkLabel(header, text=text, font=("Arial", 12, "bold"), text_color="darkorange").pack(side="left", expand=True, fill="x")
    
    scroll_table = ctk.CTkScrollableFrame(table_frame, fg_color="transparent")
    scroll_table.pack(fill="both", expand=True, padx=5, pady=(0, 15))
    scroll_container = ctk.CTkFrame(scroll_table, fg_color="transparent")
    scroll_container.pack(fill="both", expand=True)

    def update_table(query=""):
        for widget in scroll_container.winfo_children(): widget.destroy()
        filtered_list = present_details_list
        if query:
            query = query.lower()
            filtered_list = [item for item in present_details_list if query in item['name'].lower() or query in item['reg'].lower()]
        if not filtered_list:
            ctk.CTkLabel(scroll_container, text="No matching students found.", font=("Arial", 14), text_color="gray").pack(pady=50)
            return
        for item in reversed(filtered_list):
            row = ctk.CTkFrame(scroll_container, fg_color="#2c2c2c", corner_radius=8)
            row.pack(fill="x", padx=10, pady=4)
            ctk.CTkLabel(row, text=item['name'], font=("Arial", 11)).pack(side="left", expand=True, fill="x", pady=8)
            ctk.CTkLabel(row, text=item['reg'], font=("Arial", 11)).pack(side="left", expand=True, fill="x", pady=8)
            ctk.CTkLabel(row, text=item['course'], font=("Arial", 11)).pack(side="left", expand=True, fill="x", pady=8)
            ctk.CTkLabel(row, text=item.get('session', item.get('program', 'N/A')), font=("Arial", 11)).pack(side="left", expand=True, fill="x", pady=8)
            ctk.CTkLabel(row, text=item['date'], font=("Arial", 11)).pack(side="left", expand=True, fill="x", pady=8)
            ctk.CTkLabel(row, text=item['time'], font=("Arial", 11), text_color="#00c853").pack(side="left", expand=True, fill="x", pady=8)

    update_table()
    search_entry.bind("<KeyRelease>", lambda e: update_table(search_entry.get()))

# -------------------- Absent Students --------------------
def show_absent_list_content(content_area, responsive_manager, data=None):

    """Displays a list of students who have not been recognized today."""
    from tkinter import filedialog, messagebox
    from datetime import datetime

    from admin_dashboard_files import shared
    from admin_dashboard_files.shared import present_student_ids
    
    ctk.CTkLabel(content_area, text="❌ Students Absent (Live)", font=("Segoe UI", 28, "bold"), text_color="#9c27b0").pack(anchor="w", padx=30, pady=(30, 10))
    
    # Use global cache instead of DB query
    all_students = shared.all_students_cache
    absent_students = [s for s in all_students if s['reg_no'] not in present_student_ids]

    summary_bar = ctk.CTkFrame(content_area, fg_color="#1a1a1a", corner_radius=10)
    summary_bar.pack(fill="x", padx=30, pady=(0, 20))
    ctk.CTkLabel(summary_bar, text=f"Total Absent: {len(absent_students)}", font=("Arial", 14, "bold"), text_color="gray").pack(side="left", padx=20, pady=10)

    # Removed local export button as requested.

    # Absent Table
    search_frame = ctk.CTkFrame(content_area, fg_color="transparent")
    search_frame.pack(fill="x", padx=30, pady=(0, 10))
    search_entry = ctk.CTkEntry(search_frame, placeholder_text="Search Absent...", width=350, height=35, corner_radius=10)
    search_entry.pack(side="left", padx=10)

    table_frame = ctk.CTkFrame(content_area, fg_color="#1a1a1a", corner_radius=12)
    table_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
    header = ctk.CTkFrame(table_frame, fg_color="#0f0f0f", corner_radius=10)
    header.pack(fill="x", padx=15, pady=15)
    cols = ["Name", "Reg No", "Course", "Program", "Department"]
    for text in cols:
        ctk.CTkLabel(header, text=text, font=("Arial", 12, "bold"), text_color="violet").pack(side="left", expand=True, fill="x")

    scroll_table = ctk.CTkScrollableFrame(table_frame, fg_color="transparent")
    scroll_table.pack(fill="both", expand=True, padx=5, pady=(0, 15))
    scroll_container = ctk.CTkFrame(scroll_table, fg_color="transparent")
    scroll_container.pack(fill="both", expand=True)

    def update_table(query=""):
        for widget in scroll_container.winfo_children(): widget.destroy()
        filtered_list = absent_students
        if query:
            query = query.lower()
            filtered_list = [s for s in absent_students if query in s['name'].lower() or query in s['reg_no'].lower()]
        if not filtered_list:
            ctk.CTkLabel(scroll_container, text="No matching absent students found.", font=("Arial", 14), text_color="gray").pack(pady=50)
            return
        for s in reversed(filtered_list):
            row = ctk.CTkFrame(scroll_container, fg_color="#2c2c2c", corner_radius=8)
            row.pack(fill="x", padx=10, pady=4)
            for key in ['name', 'reg_no', 'course', 'program', 'department']:
                ctk.CTkLabel(row, text=s[key], font=("Arial", 11)).pack(side="left", expand=True, fill="x", pady=8)

    update_table()
    search_entry.bind("<KeyRelease>", lambda e: update_table(search_entry.get()))

# -------------------- All Students List --------------------
def show_all_students_list_content(content_area, responsive_manager, data=None):
    """Displays a complete list of all students registered in the system."""
    from tkinter import messagebox
    from db import get_student_by_regno, get_connection
    
    ctk.CTkLabel(content_area, text="👥 Registered Students Database", 
                 font=("Segoe UI", 28, "bold"), text_color="#3b9dd8").pack(anchor="w", padx=30, pady=(30, 20))

    students = data if data is not None else []
    if data is None:
        from db import get_all_students_full
        students = get_all_students_full()



    summary_bar = ctk.CTkFrame(content_area, fg_color="#1a1a1a", corner_radius=10)
    summary_bar.pack(fill="x", padx=30, pady=(0, 20))
    ctk.CTkLabel(summary_bar, text=f"Total Registered: {len(students)}", font=("Arial", 14, "bold"), text_color="gray").pack(side="left", padx=20, pady=10)


    # Search Bar
    search_frame = ctk.CTkFrame(content_area, fg_color="transparent")
    search_frame.pack(fill="x", padx=30, pady=(0, 10))
    search_entry = ctk.CTkEntry(search_frame, placeholder_text="🔍 Search Name, ID or Course...", width=450, height=35, corner_radius=10)
    search_entry.pack(side="left")

    # Details Function (Replaces content inline instead of a popup)
    def show_student_details(reg_no):
        from db import get_student_by_regno, get_student_attendance_stats
        student = get_student_by_regno(reg_no)
        if not student:
            messagebox.showerror("Error", "Student not found in database.")
            return

        # Fetch Analytics Data
        stats = get_student_attendance_stats(reg_no, student.get('course')) or {}

        # Clear current list view
        for widget in content_area.winfo_children():
            widget.destroy()

        # Fallback to reg_no if name missing
        st_name = student.get('full_name') or student.get('name') or reg_no 

        # Header with Back Button
        header_frame = ctk.CTkFrame(content_area, fg_color="transparent")
        header_frame.pack(fill="x", padx=30, pady=(30, 20))
        
        def go_back():
            for widget in content_area.winfo_children():
                widget.destroy()
            show_all_students_list_content(content_area, responsive_manager)

        ctk.CTkButton(header_frame, text="⬅ Go Back", width=100, height=35, 
                      fg_color="#333333", hover_color="#444444", text_color="white", font=("Arial", 12, "bold"),
                      command=go_back).pack(side="left")
                      
        ctk.CTkLabel(header_frame, text=f"Student Profile", 
                     font=("Segoe UI", 24, "bold"), text_color="#3b9dd8").pack(side="left", padx=20)

        # Main Details Container
        main_frame = ctk.CTkFrame(content_area, fg_color="#1a1a1a", corner_radius=15)
        main_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
        
        # Profile Info layout
        profile_top = ctk.CTkFrame(main_frame, fg_color="transparent")
        profile_top.pack(fill="x", pady=10)
        
        # Profile Picture
        img_path = student.get('image_path')
        if img_path and os.path.exists(img_path):
            try:
                img = Image.open(img_path)
                ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=(120, 120))
                ctk.CTkLabel(profile_top, image=ctk_img, text="").pack(pady=(10, 5))
            except:
                ctk.CTkLabel(profile_top, text="👤", font=("Arial", 70)).pack(pady=(10, 5))
        else:
            ctk.CTkLabel(profile_top, text="👤", font=("Arial", 70)).pack(pady=(10, 5))

        ctk.CTkLabel(profile_top, text=st_name, font=("Segoe UI", 24, "bold"), text_color="darkorange").pack(pady=5)
        
        info_split = ctk.CTkFrame(main_frame, fg_color="transparent")
        info_split.pack(fill="x", padx=40, pady=10)
        
        # Basic Details col 1
        col1 = ctk.CTkFrame(info_split, fg_color="transparent")
        col1.pack(side="left", fill="both", expand=True, padx=10)
        
        col2 = ctk.CTkFrame(info_split, fg_color="transparent")
        col2.pack(side="left", fill="both", expand=True, padx=10)
        
        fields = [
            ("Registration No", student.get('registration_no')),
            ("Email", student.get('email')),
            ("Contact Number", student.get('contact_number')),
            ("Department", student.get('department')),
            ("Course", student.get('course')),
            ("Year Level", student.get('year_level')),
            ("Session/Program", student.get('session')),
        ]

        # split fields into two columns
        mid = (len(fields) + 1) // 2
        for i, (label, val) in enumerate(fields):
            parent_col = col1 if i < mid else col2
            row = ctk.CTkFrame(parent_col, fg_color="transparent")
            row.pack(fill="x", pady=4)
            ctk.CTkLabel(row, text=f"{label}:", font=("Arial", 12, "bold"), text_color="gray", width=130, anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=str(val) if val else "N/A", font=("Arial", 12)).pack(side="left", padx=10)

        # Draw Analytics
        if stats:
            ctk.CTkLabel(main_frame, text="📊 Attendance Analytics", font=("Segoe UI", 20, "bold"), text_color="#3b9dd8").pack(anchor="w", padx=40, pady=(30, 10))
            
            stats_grid = ctk.CTkFrame(main_frame, fg_color="transparent")
            stats_grid.pack(fill="x", padx=40, pady=(0, 20))
            
            # 4 columns layout
            for c in range(4):
                stats_grid.grid_columnconfigure(c, weight=1)
                
            items = list(stats.items())
            for i, (key, val) in enumerate(items):
                r, c = i // 4, i % 4
                card = ctk.CTkFrame(stats_grid, fg_color="#2c2c2c", corner_radius=10, height=85)
                card.grid(row=r, column=c, padx=8, pady=8, sticky="nsew")
                card.pack_propagate(False)
                
                color = "white"
                text_val = val
                if isinstance(val, tuple):
                    text_val, color = val
                    
                ctk.CTkLabel(card, text=key, font=("Arial", 11, "bold"), text_color="gray").pack(pady=(15, 2))
                lbl = ctk.CTkLabel(card, text=str(text_val), font=("Arial", 16, "bold"), text_color=color)
                lbl.pack(fill="x", padx=10)
    table_frame = ctk.CTkFrame(content_area, fg_color="#1a1a1a", corner_radius=12)
    table_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
    
    header = ctk.CTkFrame(table_frame, fg_color="#0f0f0f", corner_radius=10)
    header.pack(fill="x", padx=15, pady=15)
    
    cols = ["Registration No", "Full Name", "Course", "Session", "Department", "Action"]
    for text in cols:
        ctk.CTkLabel(header, text=text, font=("Arial", 12, "bold"), text_color="darkorange").pack(side="left", expand=True, fill="x")

    scroll_table = ctk.CTkScrollableFrame(table_frame, fg_color="transparent")
    scroll_table.pack(fill="both", expand=True, padx=5, pady=(0, 15))
    scroll_container = ctk.CTkFrame(scroll_table, fg_color="transparent")
    scroll_container.pack(fill="both", expand=True)

    def update_table(query=""):
        for widget in scroll_container.winfo_children(): widget.destroy()
        
        filtered = students
        if query:
            query = query.lower()
            filtered = [s for s in students if query in s['name'].lower() or query in s['reg_no'].lower() or query in s['course'].lower()]
            
        if not filtered:
            ctk.CTkLabel(scroll_container, text="No matches found in database.", font=("Arial", 14), text_color="gray").pack(pady=50)
            return

        for s in reversed(filtered):
            row = ctk.CTkFrame(scroll_container, fg_color="#2c2c2c", corner_radius=8)
            row.pack(fill="x", padx=10, pady=4)
            # Use 'reg_no', 'name', 'course', 'session', 'department' as keys
            for key in ['reg_no', 'name', 'course', 'session', 'department']:
                ctk.CTkLabel(row, text=str(s[key]), font=("Arial", 11)).pack(side="left", expand=True, fill="x", pady=8)
                
            action_frame = ctk.CTkFrame(row, fg_color="transparent", width=80)
            action_frame.pack(side="left", expand=True, fill="x", pady=8)
            ctk.CTkButton(action_frame, text="👁 View", width=60, height=25, 
                          fg_color="#3b9dd8", hover_color="#2980b9",
                          command=lambda reg=s['reg_no']: show_student_details(reg)).pack()

    update_table()
    search_entry.bind("<KeyRelease>", lambda e: update_table(search_entry.get()))

def show_attendance_history_content(content_area, responsive_manager, data=None):
    """Displays a two-column list of archived Excel files (Present vs Absent)."""
    # Clear existing widgets
    for widget in content_area.winfo_children():  
        widget.destroy()

    from db import get_all_archives
    from admin_dashboard_files.attendance_archiver import ARCHIVE_DIR
    import os
    
    is_small = responsive_manager.is_small()
    side_pad = 10 if is_small else 30

    ctk.CTkLabel(content_area, text="📊 Attendance History (Archived)", 
                 font=("Segoe UI", 24 if is_small else 28, "bold"), text_color="#E67E22",
                 wraplength=300 if is_small else 800).pack(anchor="w", padx=side_pad, pady=(30, 5))
                 
    ctk.CTkLabel(content_area, text="Automated daily reports and summary exports.", 
                 font=("Arial", 11 if is_small else 12), text_color="gray",
                 wraplength=280 if is_small else 800).pack(anchor="w", padx=side_pad, pady=(0, 20))


    # -------------------- Summary Reports Actions --------------------

    actions_bar = ctk.CTkFrame(content_area, fg_color="#1a1a1a", corner_radius=15)
    actions_bar.pack(fill="x", padx=side_pad, pady=(0, 20))
    
    bar_side = "top" if is_small else "left"
    ctk.CTkLabel(actions_bar, text="📊 Summary Reports:", font=("Arial", 13, "bold"), text_color="gray").pack(side=bar_side, padx=20, pady=(15, 5) if is_small else 15)


    def export_summary(period):
        from datetime import datetime, timedelta
        from db import get_attendance_by_range, get_all_students_minimal
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from tkinter import filedialog, messagebox
        
        today = datetime.now()
        if period == "week":
            start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
            filename = f"Weekly_Attendance_{start}_to_{today.strftime('%Y-%m-%d')}.xlsx"
        else: # month
            start = today.replace(day=1).strftime("%Y-%m-%d")
            filename = f"Monthly_Attendance_{today.strftime('%Y-%m')}.xlsx"
            
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, title=f"Save {period.capitalize()} Report")
        if not path: return

        try:
            records = get_attendance_by_range(start, today.strftime("%Y-%m-%d"))
            all_students = get_all_students_minimal()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary Report"
            
            headers = ["Date", "Name", "Reg No", "Course", "Program", "Status", "Time In"]
            fill = PatternFill(start_color="3b9dd8", end_color="3b9dd8", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')
            
            # Map existing records for quick lookup
            # result: {(date, reg_no): status_data}
            results_map = {(str(r['date']), r['reg_no']): r for r in records}
            
            # For each day in range, list all students
            curr_row = 2
            start_dt = datetime.strptime(start, "%Y-%m-%d")
            diff_days = (today - start_dt).days + 1
            
            for i in range(diff_days):
                d_str = (start_dt + timedelta(days=i)).strftime("%Y-%m-%d")
                for s in all_students:
                    match = results_map.get((d_str, s['reg_no']))
                    status = "Present" if match else "Absent"
                    time_in = str(match['time_in']) if match else "---"
                    
                    row_data = [d_str, s['name'], s['reg_no'], s['course'], s['program'], status, time_in]
                    for col, val in enumerate(row_data, 1):
                        cell = ws.cell(row=curr_row, column=col, value=val)
                        cell.alignment = Alignment(horizontal='center')
                        if status == "Absent" and col == 6: cell.font = Font(color="FF0000")
                    curr_row += 1
            
            for col in range(1, len(headers) + 1):
                 ws.column_dimensions[chr(64+col)].width = 20
                 
            wb.save(path)
            messagebox.showinfo("Success", f"Report saved successfully!\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {e}")

    ctk.CTkButton(actions_bar, text="📥 This Week", fg_color="#3498db", hover_color="#2980b9",
                  width=180 if not is_small else 250, height=35, font=("Arial", 11 if is_small else 12, "bold"),
                  command=lambda: export_summary("week")).pack(side=bar_side, padx=10, pady=5 if is_small else 0)
                  
    ctk.CTkButton(actions_bar, text="📥 This Month", fg_color="#E67E22", hover_color="#D35400",
                  width=180 if not is_small else 250, height=35, font=("Arial", 11 if is_small else 12, "bold"),
                  command=lambda: export_summary("month")).pack(side=bar_side, padx=10, pady=(5, 15) if is_small else 0)


    # Main Container with two columns
    container = ctk.CTkFrame(content_area, fg_color="transparent")
    container.pack(fill="both", expand=True, padx=side_pad, pady=(0, 30))
    
    # Toggle between side-by-side or stacked layout
    col_side = "top" if is_small else "left"
    col_pady = (0, 15) if is_small else 0
    
    # Left Column: Present
    present_col = ctk.CTkFrame(container, fg_color="#1a1a1a", corner_radius=15)
    present_col.pack(side=col_side, fill="both", expand=True, padx=(0, 10 if not is_small else 0), pady=col_pady)
    
    header_p = ctk.CTkFrame(present_col, fg_color="transparent")
    header_p.pack(fill="x", pady=10)
    ctk.CTkLabel(header_p, text="✅ Present Records", font=("Arial", 16, "bold"), text_color="#00c853").pack(side="left", padx=20)
    
    # Right Column: Absent
    absent_col = ctk.CTkFrame(container, fg_color="#1a1a1a", corner_radius=15)
    absent_col.pack(side=col_side, fill="both", expand=True, padx=(10 if not is_small else 0, 0))
    
    header_a = ctk.CTkFrame(absent_col, fg_color="transparent")
    header_a.pack(fill="x", pady=10)
    ctk.CTkLabel(header_a, text="❌ Absent Records", font=("Arial", 16, "bold"), text_color="#9c27b0").pack(side="left", padx=20)


    def open_file(path, archive_date):
        import threading
        from admin_dashboard_files.attendance_archiver import sync_live_excel
        def _open():
            abs_path = os.path.abspath(path)
            
            # 🩹 Self-Healing: If file is missing, re-create it immediately!
            if not os.path.exists(abs_path):
                print(f"[Archiver] File missing, re-generating {archive_date}...")
                sync_live_excel(str(archive_date))
            
            try: 
                if os.path.exists(abs_path):
                    os.startfile(abs_path)
                else: 
                     raise FileNotFoundError(f"Could not regenerate file at {abs_path}")
            except Exception as e:
                from tkinter import messagebox
                messagebox.showerror("Error", f"Could not open file: {e}\n\nTIP: Please close the file in Excel if it's open.")
        threading.Thread(target=_open, daemon=True).start()

    def download_file(src_path, filename):
        import shutil
        from tkinter import filedialog, messagebox
        
        # Open standard Windows Save dialog
        dest = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=filename,
            title="Export Attendance Record"
        )
        
        if dest:
            try:
                # Direct block copy for maximum speed
                shutil.copy2(src_path, dest)
                messagebox.showinfo("Success", f"File exported successfully!\nLocation: {dest}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export file: {e}")


    def delete_single(archive_id, file_path, filename):
        from tkinter import messagebox
        from db import delete_archive_by_id
        import admin_dashboard_files.config_manager as config_manager
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this specific record?\nThis action is permanent."):
            if delete_archive_by_id(archive_id):
                # Add to Deletion Blacklist
                blacklist = config_manager.get("deleted_archives_list")
                if filename not in blacklist:
                    blacklist.append(filename)
                    config_manager.set("deleted_archives_list", blacklist)

                try:
                    if os.path.exists(file_path): os.remove(file_path)
                except: pass
                show_attendance_history_content(content_area, responsive_manager)
                messagebox.showinfo("Success", "Record deleted.")

    def clear_category(category):
        from tkinter import messagebox
        from db import delete_archives_by_category
        import admin_dashboard_files.config_manager as cfg
        if messagebox.askyesno("Confirm", f"Are you sure you want to delete ALL {category} records?\nThis cannot be undone."):
            
            # Fetch filenames before they are gone from DB?
            # Actually delete_archives_by_category returns paths. 
            # I should modify it to also return filenames or use existing archives list
            blacklist = cfg.get("deleted_archives_list")
            for a in archives:
                if a['category'] == category:
                    if a['filename'] not in blacklist:
                        blacklist.append(a['filename'])
            cfg.set("deleted_archives_list", blacklist)

            paths = delete_archives_by_category(category)
            for p in paths:
                try:
                    if os.path.exists(p): os.remove(p)
                except: pass
            show_attendance_history_content(content_area, responsive_manager)
            messagebox.showinfo("Success", f"All {category} records cleared.")

    def restore_category(category):
        from tkinter import messagebox
        from datetime import datetime
        import admin_dashboard_files.config_manager as cfg
        from admin_dashboard_files.attendance_archiver import sync_live_excel
        
        today = datetime.now().strftime("%Y-%m-%d")
        filename = f"{category.lower()}_{today}.xlsx"
        
        # 🔓 Unlock the file from the blacklist
        blacklist = cfg.get("deleted_archives_list")
        if filename in blacklist:
            blacklist.remove(filename)
            cfg.set("deleted_archives_list", blacklist)
            
        # ⚡ Re-Sync immediately
        if sync_live_excel(today):
            show_attendance_history_content(content_area, responsive_manager)
            messagebox.showinfo("Success", f"Today's {category} records have been restored!")
        else:
            messagebox.showerror("Error", "Could not restore records. Check database connection.")

    # Header Buttons: Stack below title on mobile for full visibility
    if is_small:
        p_btn_frame = ctk.CTkFrame(present_col, fg_color="transparent")
        p_btn_frame.pack(fill="x", padx=20, pady=(0, 10))
        ctk.CTkButton(p_btn_frame, text="🔄 Restore Today", height=30, fg_color="#3498db", hover_color="#2980b9",
                      command=lambda: restore_category("Present")).pack(side="left", expand=True, fill="x", padx=5)
        ctk.CTkButton(p_btn_frame, text="🗑️ Clear All", height=30, fg_color="#e74c3c", hover_color="#c0392b",
                      command=lambda: clear_category("Present")).pack(side="left", expand=True, fill="x", padx=5)

        a_btn_frame = ctk.CTkFrame(absent_col, fg_color="transparent")
        a_btn_frame.pack(fill="x", padx=20, pady=(0, 10))
        ctk.CTkButton(a_btn_frame, text="🔄 Restore Today", height=30, fg_color="#3498db", hover_color="#2980b9",
                      command=lambda: restore_category("Absent")).pack(side="left", expand=True, fill="x", padx=5)
        ctk.CTkButton(a_btn_frame, text="🗑️ Clear All", height=30, fg_color="#e74c3c", hover_color="#c0392b", 
                      command=lambda: clear_category("Absent")).pack(side="left", expand=True, fill="x", padx=5)
    else:
        p_btn_frame = ctk.CTkFrame(header_p, fg_color="transparent")
        p_btn_frame.pack(side="right", padx=10)
        ctk.CTkButton(p_btn_frame, text="🔄 Restore Today", width=110, height=25, fg_color="#3498db", hover_color="#2980b9",
                      command=lambda: restore_category("Present")).pack(side="left", padx=5)
        ctk.CTkButton(p_btn_frame, text="🗑️ Clear All", width=80, height=25, fg_color="#e74c3c", hover_color="#c0392b",
                      command=lambda: clear_category("Present")).pack(side="left", padx=5)

        a_btn_frame = ctk.CTkFrame(header_a, fg_color="transparent")
        a_btn_frame.pack(side="right", padx=10)
        ctk.CTkButton(a_btn_frame, text="🔄 Restore Today", width=110, height=25, fg_color="#3498db", hover_color="#2980b9",
                      command=lambda: restore_category("Absent")).pack(side="left", padx=5)
        ctk.CTkButton(a_btn_frame, text="🗑️ Clear All", width=80, height=25, fg_color="#e74c3c", hover_color="#c0392b", 
                      command=lambda: clear_category("Absent")).pack(side="left", padx=5)

    def create_item(parent, archive, color="#333333"):
        is_small = responsive_manager.is_small()
        item = ctk.CTkFrame(parent, fg_color="#2c2c2c", corner_radius=12)
        item.pack(fill="x", padx=10 if is_small else 15, pady=5)
        
        display_name = archive['filename'].replace(".xlsx", "")
        
        if is_small:
            # 📱 MOBILE STACK: Vertical Layout for 100% Clarity
            lbl = ctk.CTkLabel(item, text=f"📊 {display_name}", 
                         font=("Arial", 11, "bold"), anchor="w")
            lbl.pack(fill="x", padx=15, pady=(12, 5))
            
            btns = ctk.CTkFrame(item, fg_color="transparent")
            btns.pack(fill="x", padx=10, pady=(0, 10))
            
            # --- OPEN BUTTON ---
            btn_open = ctk.CTkButton(btns, text="📂 Open", height=32,
                                     fg_color=color, hover_color="#444444", font=("Arial", 10, "bold"),
                                     command=lambda p=archive['file_path'], d=archive['date']: open_file(p, d))
            btn_open.pack(side="left", expand=True, fill="x", padx=2)
            
            # --- DOWNLOAD BUTTON ---
            btn_dnld = ctk.CTkButton(btns, text="📥 Dnld", height=32,
                                     fg_color="#3d3d3d", hover_color="#4d4d4d", font=("Arial", 10, "bold"),
                                     command=lambda p=archive['file_path'], f=archive['filename']: download_file(p, f))
            btn_dnld.pack(side="left", expand=True, fill="x", padx=2)
            
            # --- DELETE BUTTON ---
            btn_del = ctk.CTkButton(btns, text="🗑️", width=40, height=32,
                                    fg_color="#c0392b", hover_color="#e74c3c",
                                    command=lambda p=archive['file_path'], aid=archive['id'], fn=archive['filename']: delete_single(aid, p, fn))
            btn_del.pack(side="left", padx=2)
        else:
            # 🖥️ DESKTOP: Traditional Horizontal Layout
            label = ctk.CTkLabel(item, text=f"📊 {display_name}", 
                                 font=("Arial", 12, "bold"), anchor="w")
            label.pack(side="left", padx=15, pady=12)
            
            btns = ctk.CTkFrame(item, fg_color="transparent")
            btns.pack(side="right", padx=10)
            
            # --- OPEN BUTTON ---
            btn_open = ctk.CTkButton(btns, text="📂 Open", width=70, height=28,
                                     fg_color=color, hover_color="#444444", text_color="white", font=("Arial", 11),
                                     command=lambda p=archive['file_path'], d=archive['date']: open_file(p, d))
            btn_open.pack(side="left", padx=5)
            
            # --- DOWNLOAD BUTTON ---
            btn_dnld = ctk.CTkButton(btns, text="📥 Dnld", width=70, height=28,
                                     fg_color="#333333", hover_color="#444444", text_color="white", font=("Arial", 11),
                                     command=lambda p=archive['file_path'], f=archive['filename']: download_file(p, f))
            btn_dnld.pack(side="left", padx=5)
            
            # --- DELETE BUTTON ---
            btn_del = ctk.CTkButton(btns, text="🗑️", width=30, height=28,
                                    fg_color="#c0392b", hover_color="#e74c3c", text_color="white",
                                    command=lambda p=archive['file_path'], aid=archive['id'], fn=archive['filename']: delete_single(aid, p, fn))
            btn_del.pack(side="left", padx=5)


    # Fetch from DB
    archives = data if data is not None else get_all_archives()

    
    # Scrollable Areas
    present_scroll = ctk.CTkScrollableFrame(present_col, fg_color="transparent")
    present_scroll.pack(fill="both", expand=True, padx=5, pady=5)
    
    absent_scroll = ctk.CTkScrollableFrame(absent_col, fg_color="transparent")
    absent_scroll.pack(fill="both", expand=True, padx=5, pady=5)

    if not archives:
        ctk.CTkLabel(present_scroll, text="No archives recorded.", text_color="gray").pack(pady=20)
        ctk.CTkLabel(absent_scroll, text="No archives recorded.", text_color="gray").pack(pady=20)
    else:
        for a in archives:
            if a['category'] == "Present":
                create_item(present_scroll, a, "#00c853")
            elif a['category'] == "Absent":
                create_item(absent_scroll, a, "#9c27b0")

