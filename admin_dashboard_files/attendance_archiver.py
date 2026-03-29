import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import admin_dashboard_files.config_manager as config_manager
from db import get_attendance_by_date, get_all_students_minimal, record_archive_entry

ARCHIVE_DIR = "attendance_history"

def check_and_archive_attendance():
    """
    Checks if the day has changed since the last run.
    If so, it finalizes the previous day's records and switches to the new date.
    """
    last_date = config_manager.get("last_archive_date")
    today = datetime.now().strftime("%Y-%m-%d")
    
    if last_date == today: return # Date hasn't changed
        
    if not last_date:
        config_manager.set("last_archive_date", today)
        return

    # Finalize previous day(s) one last time to ensure they are 100% accurate
    print(f"[Archiver] Date change: Finalizing archives for {last_date}...")
    sync_live_excel(last_date)
    
    # Switch to new date tracker
    config_manager.set("last_archive_date", today)

def sync_live_excel(date_str):
    """
    Syncs the database state for the given date into Excel files on disk.
    Updates existing files with new attendance and records them in the DB.
    """
    try:
        os.makedirs(ARCHIVE_DIR, exist_ok=True)
        
        # 1. Fetch Present and Absent records
        present_records = get_attendance_by_date(date_str)
        all_students = get_all_students_minimal()
        
        if not all_students:
            print(f"[Archiver] No students in DB, skipping sync for {date_str}.")
            return
            
        present_ids = {r['reg_no'] for r in present_records}
        absent_records = [s for s in all_students if s['reg_no'] not in present_ids]
        
        # 2. present Excel (Live)
        present_filename = f"present_{date_str}.xlsx"
        blacklist = config_manager.get("deleted_archives_list")
        if present_filename not in blacklist:
            p_path = os.path.join(ARCHIVE_DIR, present_filename)
            _create_excel(
                p_path, "Present Students",
                ["Name", "Reg No", "Course", "Program", "Time In"],
                [[r['name'], r['reg_no'], r['course'], r['program'], str(r.get('time_in', 'N/A'))] for r in present_records],
                "00C853" # Green
            )
            _ensure_db_record(present_filename, date_str, "Present", p_path)

        # 3. absent Excel (Live)
        absent_filename = f"absent_{date_str}.xlsx"
        if absent_filename not in blacklist:
            a_path = os.path.join(ARCHIVE_DIR, absent_filename)
            _create_excel(
                a_path, "Absent Students",
                ["Name", "Reg No", "Course", "Program", "Department"],
                [[s['name'], s['reg_no'], s['course'], s['program'], s['department']] for s in absent_records],
                "9C27B0" # Purple
            )
            _ensure_db_record(absent_filename, date_str, "Absent", a_path)
        return True
    except Exception as e:
        print(f"[Archiver] Error in sync_live_excel ({date_str}): {e}")
        return False


def _ensure_db_record(filename, date_str, category, path):
    """
    Connects to the archive table and ensures this file is recorded.
    Now uses UNIQUE constraints in DB for safety.
    """
    from db import record_archive_entry
    record_archive_entry(filename, date_str, category, path)


def _create_excel(path, title, headers, rows, color_hex):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border
        
    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = alignment
            
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64+col)].width = 25
        
    try:
        wb.save(path)
        return True
    except PermissionError:
        print(f"[Archiver] Update Blocked: '{path}' is open in Excel. Please close it to sync attendance.")
        return False
    except Exception as e:
        print(f"[Archiver] Failed to save Excel: {e}")
        return False
